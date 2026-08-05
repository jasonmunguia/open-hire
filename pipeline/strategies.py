"""Work out where one candidate's resume ends and the next begins.

This is the hard part of the whole project, and the part most likely to be
wrong on a PDF nobody has seen before. Four strategies, tried in order of how
much they can be trusted:

  1. bookmarks  — the PDF's own outline marks each resume. Exact.
  2. roster     — a spreadsheet carries an explicit start-page column. Exact.
  3. handshake  — a roster table is printed in the PDF's first pages, and the
                  resumes follow in the same order. ~98% on real exports.
  4. heuristic  — no structure at all; guess from page layout. Unreliable, and
                  it says so: every candidate comes back flagged.

Anything a strategy is unsure about is flagged rather than silently trusted.
A wrong boundary means a reviewer judges someone by a stranger's resume, which
is far worse than being told to check 8 of them by hand.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from itertools import pairwise

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")
STRICT_EMAIL_RE = re.compile(r"^[\w.+-]+@[\w-]+\.[\w.]+$")
WORD_RE = re.compile(r"[a-z]{2,}")
HANDSHAKE_MARKER = "Job applicants as of"
LONG_RESUME_PAGES = 6
NAME_STOPWORDS = {
    "university", "college", "school", "the", "state", "institute", "and",
    "masters", "alumni", "senior", "junior", "resume", "curriculum", "vitae",
}


@dataclass
class Candidate:
    """One person plus the page range their resume occupies (end exclusive)."""
    name: str
    start: int
    end: int
    email: str | None = None
    school: str | None = None
    school_year: str | None = None
    majors: str | None = None
    grad_date: str | None = None
    external_id: str | None = None
    flags: list[str] = field(default_factory=list)

    @property
    def pages(self) -> int:
        return self.end - self.start


def _tokens(text: str) -> set[str]:
    return set(WORD_RE.findall(text.lower()))


def _name_parts(name: str) -> list[str]:
    return [t.lower() for t in re.split(r"[^A-Za-z]+", name or "")
            if len(t) > 1 and t.lower() not in NAME_STOPWORDS]


def _close_ranges(people: list[Candidate], page_count: int) -> list[Candidate]:
    """Each resume runs until the next one starts."""
    people.sort(key=lambda c: c.start)
    for a, b in pairwise(people):
        a.end = b.start
    if people:
        people[-1].end = page_count
    for c in people:
        if c.pages > LONG_RESUME_PAGES:
            c.flags.append("long")
        if c.pages < 1:
            c.flags.append("empty")
    return people


# ── 1. PDF bookmarks ────────────────────────────────────────────────────────

def detect_bookmarks(doc) -> bool:
    return len(doc.get_toc()) >= 2


def from_bookmarks(doc) -> list[Candidate]:
    """Bookmark title is the name; its page is the start.

    Titles are often prefixed like "AG #0001 — Jane Doe" or "12. Jane Doe".
    The prefix becomes the external id and the remainder the name.
    """
    people = []
    for _level, title, page in doc.get_toc():
        raw = (title or "").strip()
        external_id, name = None, raw
        for sep in ("—", " - ", " – ", "|", ":"):
            if sep in raw:
                left, right = raw.split(sep, 1)
                if len(left.strip()) <= 24:
                    external_id, name = left.strip(), right.strip()
                break
        else:
            m = re.match(r"^\s*(\d{1,5})[.)]\s+(.*)$", raw)
            if m:
                external_id, name = m.group(1), m.group(2).strip()
        people.append(Candidate(name=name or raw, start=page - 1, end=page,
                                external_id=external_id))
    return _close_ranges(people, doc.page_count)


# ── 2. Roster spreadsheet with an explicit start page ───────────────────────

ROSTER_ALIASES = {
    "name": ("name", "candidate name", "full name", "applicant"),
    "email": ("email", "email address", "e-mail"),
    "school": ("school", "university", "institution"),
    "school_year": ("school year", "year", "class year", "level"),
    "majors": ("majors", "major", "degree", "field of study"),
    "grad_date": ("graduation date", "grad date", "graduation"),
    "start": ("pdf start page", "start page", "page", "pdf page"),
    "pages": ("resume pages", "pages", "page count", "num pages"),
    "external_id": ("candidate id", "id", "applicant id"),
}


def _match_header(cell: str) -> str | None:
    key = str(cell or "").strip().lower()
    for field_name, aliases in ROSTER_ALIASES.items():
        if key in aliases:
            return field_name
    return None


def read_roster_rows(path: str, sheet: str | None = None) -> list[dict]:
    """Read an .xlsx or .csv roster. Returns dicts keyed by our field names."""
    if path.lower().endswith(".csv"):
        import csv
        with open(path, newline="", encoding="utf-8-sig") as fh:
            rows = [list(r) for r in csv.reader(fh)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

    header_i, columns = None, {}
    for i, row in enumerate(rows[:40]):
        mapped = {}
        for j, cell in enumerate(row or []):
            field_name = _match_header(cell)
            if field_name and field_name not in mapped:
                mapped[field_name] = j
        if "name" in mapped and len(mapped) >= 2:
            header_i, columns = i, mapped
            break
    if header_i is None:
        raise ValueError(
            f"No header row found in {path}. Expected a row containing at least "
            f"a 'Name' column. Recognised headers: "
            + ", ".join(sorted({a for v in ROSTER_ALIASES.values() for a in v}))
        )

    out = []
    for row in rows[header_i + 1:]:
        if not row:
            continue
        rec = {}
        for field_name, j in columns.items():
            val = row[j] if j < len(row) else None
            rec[field_name] = None if val is None else str(val).strip()
        if rec.get("name"):
            out.append(rec)
    return out


def detect_roster(roster_rows: list[dict] | None) -> bool:
    return bool(roster_rows) and any(r.get("start") for r in roster_rows)


def from_roster(doc, roster_rows: list[dict]) -> list[Candidate]:
    people = []
    for r in roster_rows:
        if not r.get("start"):
            continue
        try:
            start = int(float(r["start"])) - 1        # spreadsheets are 1-based
            span = int(float(r.get("pages") or 1))
        except (TypeError, ValueError):
            continue
        people.append(Candidate(
            name=r["name"], start=start, end=start + max(span, 1),
            email=(r.get("email") or "").lower() or None,
            school=r.get("school"), school_year=r.get("school_year"),
            majors=r.get("majors"), grad_date=(r.get("grad_date") or "")[:10] or None,
            external_id=r.get("external_id"),
        ))
    people.sort(key=lambda c: c.start)
    for c in people:
        if c.start < 0 or c.end > doc.page_count or c.end <= c.start:
            c.flags.append("bad-range")
        if c.pages > LONG_RESUME_PAGES:
            c.flags.append("long")
    return people


# ── 3. Handshake-style roster table printed inside the PDF ──────────────────

def detect_handshake(doc) -> bool:
    return HANDSHAKE_MARKER in doc[0].get_text()[:250]


def _derive_bands(words):
    """Column x-positions differ page to page because the exporter auto-fits
    column widths. Always read them from the page's own header row."""
    seen, school_xs = {}, []
    for w in sorted(words, key=lambda w: (w[1], w[0])):
        if w[4] == "School":
            school_xs.append(w[0])
        elif w[4] == "Name" and "name" not in seen:
            seen["name"] = w[0]
        elif w[4] == "Email" and "email" not in seen:
            seen["email"] = w[0]
        elif w[4] == "Major" and "majors" not in seen:
            seen["majors"] = w[0]
    if len(school_xs) >= 2:
        seen["school"], seen["school_year"] = school_xs[0], school_xs[1]
    elif school_xs:
        seen["school"] = school_xs[0]
    if len(seen) < 4:
        return None, None
    ordered = sorted(seen.items(), key=lambda kv: kv[1])
    bands = {}
    for i, (col, x) in enumerate(ordered):
        nxt = ordered[i + 1][1] if i + 1 < len(ordered) else float("inf")
        bands[col] = (x - 3, nxt - 3)
    header_y = max(w[1] for w in words if w[4] in ("Name", "Major"))
    return bands, header_y


def parse_handshake_roster(doc):
    """Returns (rows, index of the first resume page)."""
    rows, first_resume = [], doc.page_count
    for page_no in range(doc.page_count):
        if HANDSHAKE_MARKER not in doc[page_no].get_text()[:250]:
            first_resume = page_no
            break
        words = doc[page_no].get_text("words")
        bands, header_y = _derive_bands(words)
        if not bands:
            continue
        ex0, ex1 = bands["email"]
        anchors = sorted(
            (w for w in words
             if ex0 <= w[0] <= ex1 and STRICT_EMAIL_RE.match(w[4]) and w[1] > header_y),
            key=lambda w: w[1])
        for i, anchor in enumerate(anchors):
            top = anchor[1] - 6
            bottom = anchors[i + 1][1] - 6 if i + 1 < len(anchors) else float("inf")
            cells = {col: [] for col in bands}
            for w in sorted(words, key=lambda w: (w[1], w[0])):
                if w[1] <= header_y or not (top <= w[1] < bottom):
                    continue
                for col, (x0, x1) in bands.items():
                    if x0 <= w[0] < x1:
                        cells[col].append(w[4])
                        break
            rec = {col: " ".join(v).strip() for col, v in cells.items()}
            rec["email"] = anchor[4].lower()
            rows.append(rec)
    return rows, first_resume


def from_handshake(doc):
    """Resumes follow the roster in the same order, so walk forward and never
    look back. That monotonicity is what lets an unmatched candidate still be
    recovered from the gap between their matched neighbours."""
    rows, first_resume = parse_handshake_roster(doc)
    if not rows:
        return []

    prepared = []
    for r in rows:
        parts = _name_parts(r.get("name", ""))
        prepared.append({
            "email": r["email"],
            "local": r["email"].split("@")[0],
            "first": parts[0] if parts else "",
            "last": parts[-1] if parts else "",
            "tokens": set(parts),
        })

    signals = []
    for p in range(first_resume, doc.page_count):
        text = doc[p].get_text()
        signals.append((p, _tokens(text[:450]),
                        {e.lower() for e in EMAIL_RE.findall(text)}))

    starts, cursor = {}, 0
    for page_no, head, emails in signals:
        best_score, best_i = 0, None
        for i in range(cursor, min(cursor + 6, len(prepared))):
            c, score = prepared[i], 0
            # Exact email, or same local-part on a different domain — people
            # often list a personal address on the resume and a school one on
            # the application.
            if c["email"] in emails or (
                c["local"] and any(c["local"] == e.split("@")[0] for e in emails)
            ):
                score += 3
            if c["last"] and c["last"] in head:
                score += 2
            if c["first"] and c["first"] in head:
                score += 2
            if c["tokens"] & head:
                score += 1
            if score > best_score:
                best_score, best_i = score, i
        if best_score >= 3 and best_i not in starts:
            starts[best_i] = page_no
            cursor = best_i + 1

    people = []
    for i, r in enumerate(rows):
        people.append(Candidate(
            name=r.get("name", ""), start=starts.get(i, -1), end=-1,
            email=r["email"], school=r.get("school") or None,
            school_year=r.get("school_year") or None, majors=r.get("majors") or None,
            flags=[] if i in starts else ["unmatched"],
        ))

    # Unmatched people inherit the next matched start, so their pages land in
    # the gap rather than being dropped entirely.
    next_start = doc.page_count
    for c in reversed(people):
        if c.start < 0:
            c.start = next_start
        else:
            next_start = c.start
    return _close_ranges(people, doc.page_count)


# ── 4. Heuristic, for a PDF with no structure at all ────────────────────────

def from_heuristic(doc) -> list[Candidate]:
    """A new resume usually starts with a large-text line near the top of the
    page followed by contact details. This is a guess and is labelled as one:
    every candidate is flagged `heuristic` for human confirmation.
    """
    people = []
    for p in range(doc.page_count):
        page = doc[p]
        blocks = page.get_text("dict")["blocks"]
        biggest, biggest_text = 0.0, ""
        for b in blocks:
            for line in b.get("lines", []):
                y = line["bbox"][1]
                if y > page.rect.height * 0.35:
                    continue
                for span in line["spans"]:
                    if span["size"] > biggest and span["text"].strip():
                        biggest, biggest_text = span["size"], span["text"].strip()
        head = page.get_text()[:500]
        looks_like_start = (
            biggest >= 13
            and 2 <= len(biggest_text.split()) <= 6
            and bool(EMAIL_RE.search(head))
        )
        if looks_like_start or p == 0:
            people.append(Candidate(
                name=biggest_text or f"Unknown (page {p + 1})",
                start=p, end=p + 1,
                email=(EMAIL_RE.search(head).group(0).lower()
                       if EMAIL_RE.search(head) else None),
                flags=["heuristic"],
            ))
    return _close_ranges(people, doc.page_count)


# ── dispatch ────────────────────────────────────────────────────────────────

def choose(doc, roster_rows: list[dict] | None, forced: str | None = None):
    """Returns (strategy_name, candidates)."""
    if forced:
        table = {
            "bookmarks": lambda: from_bookmarks(doc),
            "roster": lambda: from_roster(doc, roster_rows or []),
            "handshake": lambda: from_handshake(doc),
            "heuristic": lambda: from_heuristic(doc),
        }
        if forced not in table:
            raise ValueError(f"Unknown strategy {forced!r}. "
                             f"Choose from: {', '.join(table)}")
        return forced, table[forced]()

    if detect_roster(roster_rows):
        return "roster", from_roster(doc, roster_rows)
    if detect_bookmarks(doc):
        return "bookmarks", from_bookmarks(doc)
    if detect_handshake(doc):
        return "handshake", from_handshake(doc)
    return "heuristic", from_heuristic(doc)


def enrich_from_roster(people: list[Candidate], roster_rows: list[dict]) -> int:
    """Fill blank metadata by matching a roster on email, then on name.

    Useful when boundaries came from bookmarks (exact) but the school, year and
    majors live in a separate spreadsheet.
    """
    by_email, by_name = {}, {}
    for r in roster_rows:
        if r.get("email"):
            by_email[r["email"].lower()] = r
        if r.get("name"):
            by_name[" ".join(_name_parts(r["name"]))] = r
    filled = 0
    for c in people:
        r = (by_email.get((c.email or "").lower())
             or by_name.get(" ".join(_name_parts(c.name))))
        if not r:
            continue
        before = (c.email, c.school, c.school_year, c.majors, c.grad_date)
        c.email = c.email or (r.get("email") or "").lower() or None
        c.school = c.school or r.get("school")
        c.school_year = c.school_year or r.get("school_year")
        c.majors = c.majors or r.get("majors")
        c.grad_date = c.grad_date or (r.get("grad_date") or "")[:10] or None
        if (c.email, c.school, c.school_year, c.majors, c.grad_date) != before:
            filled += 1
    return filled
